# -*- coding: utf-8 -*-
"""
Actualiza las planillas de analisis con datos de los manuales de fabricante:
- MANTENIMIENTO EN ANALISIS.xlsx (hoja maintenance_plans): corrige intervalos
  y agrega tareas recomendadas para MP (Volvo D16), CR (Twin Disc MG-5170DC)
  y MA (Cummins serie ES/C).
- REPUESTOS ANALISIS.xlsx (hoja spares): agrega repuestos de caja Twin Disc
  y generadores Cummins, y completa P/N del filtro Twin Disc.
Hace backup de ambos archivos antes de tocarlos.
"""
import openpyxl, shutil, datetime, sys, re

io = sys.stdout.buffer
MANT = r"C:\CMS3.0\MisDocs\MANTENIMIENTO EN ANALISIS.xlsx"
REP  = r"C:\CMS3.0\MisDocs\REPUESTOS ANALISIS.xlsx"
stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")

# ---------- backups ----------
for p in (MANT, REP):
    bak = p.replace(".xlsx", f".bak-{stamp}.xlsx")
    shutil.copyfile(p, bak)
    io.write(f"backup -> {bak}\n".encode("utf-8"))

def C(L): return openpyxl.utils.column_index_from_string(L)

# referencia por activo (sfiGroup, sfiCode, assetId, assetName literal, triggerResultMode)
ASSET = {
 "M01-MP-BR": dict(sg=6, sf=600, aid="cmqhe9xy8000n28l4dad0rnkl", an="Motor Principal Babor",   trm="DUE_ONLY"),
 "M01-MP-ER": dict(sg=6, sf=600, aid="cmqhe9xv3000028l4tv5jlqs2", an="Motor Principal Estribor", trm="DUE_ONLY"),
 "M01-CR-BR": dict(sg=6, sf=600, aid="cmqhe9y00001g28l49sv413w3", an="Caja reductora Babor",     trm="AUTO_WO"),
 "M01-CR-ER": dict(sg=6, sf=600, aid="cmqhe9xzo001a28l4orwnvw4f", an="Caja reductora Estribor",   trm="AUTO_WO"),
 "M01-MA-BR": dict(sg=6, sf=600, aid="cmqhe9y20002g28l4beoq97tn", an="Motor Auxiliar Babor",      trm="DUE_ONLY"),
 "M01-MA-ER": dict(sg=6, sf=600, aid="cmqhe9y10001u28l4tuahxdbd", an="Motor Auxiliar Estribor",   trm="DUE_ONLY"),
}
TENANT = "cmqhafl85006ysll4f0hecqcu"

# =========================================================
# 1) MANTENIMIENTO
# =========================================================
wb = openpyxl.load_workbook(MANT)
ws = wb["maintenance_plans"]

# --- 1a) CORRECCIONES sobre filas existentes ---
# clave (assetCode, taskCode) -> dict de cambios por letra de columna
CORR = {
 # Volvo MP: intervalos al manual
 ("M01-MP-BR","M01-MP-BR-03"): {"H":"Cambio de impeller / rotor de bomba de agua de mar","M":1000},
 ("M01-MP-ER","M01-MP-ER-03"): {"H":"Cambio de impeller / rotor de bomba de agua de mar","M":1000},
 ("M01-MP-BR","M01-MP-BR-04"): {"M":1000},
 ("M01-MP-ER","M01-MP-ER-04"): {"M":1000},
 ("M01-MP-BR","M01-MP-BR-05"): {"M":2000},
 ("M01-MP-ER","M01-MP-ER-05"): {"M":2000},
 ("M01-MP-BR","M01-MP-BR-07"): {"K":"HOURS","L":None,"M":1000},
 ("M01-MP-ER","M01-MP-ER-07"): {"K":"HOURS","L":None,"M":1000},
 ("M01-MP-BR","M01-MP-BR-08"): {"K":"HOURS","L":None,"M":8000},
 ("M01-MP-ER","M01-MP-ER-08"): {"K":"HOURS","L":None,"M":8000},
 # Cummins MA: reemplazar tareas de motor grande por esquema ES/C
 ("M01-MA-BR","M01-MA-BR-18"): {"H":"Calibracion de inyectores y bomba inyectora de combustible (Servicio Tecnico Bosch) - Rev E",
                                "I":"Segun manual Cummins serie ES/C - Revision E (cada 2900-3000 h / 24 meses).",
                                "J":"MAINTENANCE","K":"HOURS","L":None,"M":3000},
 ("M01-MA-ER","M01-MA-ER-15"): {"H":"Calibracion de inyectores y bomba inyectora de combustible (Servicio Tecnico Bosch) - Rev E",
                                "I":"Segun manual Cummins serie ES/C - Revision E (cada 2900-3000 h / 24 meses).",
                                "J":"MAINTENANCE","K":"HOURS","L":None,"M":3000},
 ("M01-MA-BR","M01-MA-BR-17"): {"H":"Medicion de gases del carter (blow-by) - Rev E",
                                "I":"Segun manual Cummins serie ES/C - Revision E (cada 2900-3000 h / 24 meses).",
                                "J":"INSPECTION","K":"HOURS","L":None,"M":3000},
 ("M01-MA-ER","M01-MA-ER-17"): {"H":"Medicion de gases del carter (blow-by) - Rev E",
                                "I":"Segun manual Cummins serie ES/C - Revision E (cada 2900-3000 h / 24 meses).",
                                "J":"INSPECTION","K":"HOURS","L":None,"M":3000},
}
applied = 0
for r in range(2, ws.max_row+1):
    ac = ws.cell(row=r, column=C("B")).value
    tc = ws.cell(row=r, column=C("G")).value
    if (ac, tc) in CORR:
        for L, v in CORR[(ac, tc)].items():
            ws.cell(row=r, column=C(L)).value = v
        applied += 1
io.write(f"correcciones aplicadas: {applied}/{len(CORR)}\n".encode("utf-8"))

# --- 1b) NUEVAS tareas ---
# (assetCode, suffix, title, taskType, triggerType, months, hours, estHours, desc)
NEW = [
 # ---- Volvo MP (ambas bandas) ----
 *[(a, 22, "Cambio de correas de transmision","MAINTENANCE","HOURS",None,2000,2,
    "Volvo D16 - Servicio C (cada 2000 h).") for a in ("M01-MP-BR","M01-MP-ER")],
 *[(a, 23, "Cambio de anodos del intercambiador de calor / enfriador de aire de carga","MAINTENANCE","HOURS",None,1000,2,
    "Volvo D16 - Servicio B (cada 1000 h). Los anodos no deben pintarse.") for a in ("M01-MP-BR","M01-MP-ER")],
 *[(a, 24, "Cambio del kit de desgaste (wear kit) de la bomba de agua de mar","MAINTENANCE","HOURS",None,3000,8,
    "Volvo D16 - Servicio D (cada 3000 h).") for a in ("M01-MP-BR","M01-MP-ER")],
 *[(a, 25, "Limpieza del filtro de agua de mar","INSPECTION","HOURS",None,500,1,
    "Volvo D16 - Servicio A (cada 500 h).") for a in ("M01-MP-BR","M01-MP-ER")],
 # ---- Twin Disc CR (ambas bandas) ----
 *[(a, 6, "Engrase de sellos del eje de salida","MAINTENANCE","HOURS",None,100,1,
    "Twin Disc MG-5170DC - cada 100 h o al amarrar. Grasa de litio NLGI N2 (tipo bomba de agua).") for a in ("M01-CR-BR","M01-CR-ER")],
 *[(a, 7, "Inspeccion del acople torsional flexible","INSPECTION","HOURS",None,2000,2,
    "Twin Disc MG-5170DC - primera a 100 h, luego cada 2000 h o 6 meses. Buscar grietas, separacion del caucho y juego en el buje.") for a in ("M01-CR-BR","M01-CR-ER")],
 *[(a, 8, "Control / cambio de anodos de zinc del intercambiador de calor","INSPECTION","MONTHS",3,None,1,
    "Twin Disc MG-5170DC - cada 90 dias. Reemplazar si esta consumido mas del 50%.") for a in ("M01-CR-BR","M01-CR-ER")],
 *[(a, 9, "Inspeccion visual periodica: montajes, mangueras, sellos y fugas de aceite","INSPECTION","MONTHS",6,None,1,
    "Twin Disc MG-5170DC - inspeccion periodica de lineas de aceite, sellos de eje y montajes.") for a in ("M01-CR-BR","M01-CR-ER")],
 # ---- Cummins MA (ambas bandas) ----
 *[(a, 22, "Limpieza del alternador con aire comprimido seco (28 PSI) y verificacion de conectores","INSPECTION","HOURS",None,250,1,
    "Cummins ES/C - Revision B (cada 225-250 h / 3 meses).") for a in ("M01-MA-BR","M01-MA-ER")],
 *[(a, 23, "Drenaje del separador de agua/combustible y revision de banjos","MAINTENANCE","HOURS",None,250,1,
    "Cummins ES/C - Revision B. Sustituir la arandela de cobre si hay fugas en los banjos.") for a in ("M01-MA-BR","M01-MA-ER")],
 *[(a, 24, "Medicion de contrapresion de gases de escape","INSPECTION","HOURS",None,500,1,
    "Cummins ES/C - Revision C (cada 450-500 h / 6 meses).") for a in ("M01-MA-BR","M01-MA-ER")],
 *[(a, 25, "Limpieza de bornes de bateria, densidad de electrolito y verificacion de cargador (14 V)","INSPECTION","HOURS",None,500,1,
    "Cummins ES/C - Revision C (cada 450-500 h / 6 meses).") for a in ("M01-MA-BR","M01-MA-ER")],
 *[(a, 26, "Cambio de correa del ventilador","MAINTENANCE","HOURS",None,1500,2,
    "Cummins ES/C - Revision D (cada 1450-1500 h / 12 meses).") for a in ("M01-MA-BR","M01-MA-ER")],
 *[(a, 27, "Medicion de resistencias de estator y rotor del alternador (principal y de excitacion)","INSPECTION","HOURS",None,3000,2,
    "Cummins ES/C - Revision E (cada 2900-3000 h / 24 meses).") for a in ("M01-MA-BR","M01-MA-ER")],
]

row = ws.max_row + 1
added = 0
for (ac, suf, title, ttype, trig, months, hours, est, desc) in NEW:
    a = ASSET[ac]
    tc = f"{ac}-{suf:02d}"
    vals = {
      "A":"M01","B":ac,"C":a["sg"],"D":a["sf"],"E":a["aid"],"F":a["an"],
      "G":tc,"H":title,"I":desc,"J":ttype,"K":trig,"L":months,"M":hours,
      "N":est,"O":"Jefe de Maquinas","V":a["trm"],"W":"AUTO","AG":"ACTIVE","AT":TENANT,
    }
    for L, v in vals.items():
        ws.cell(row=row, column=C(L)).value = v
    row += 1; added += 1
io.write(f"tareas nuevas agregadas: {added}\n".encode("utf-8"))
wb.save(MANT)
io.write(b"MANTENIMIENTO guardado.\n")

# =========================================================
# 2) REPUESTOS
# =========================================================
wb2 = openpyxl.load_workbook(REP)
ws2 = wb2["spares"]

# (sku, name, category, criticality, manufacturer, model, mfgPN, unit,
#  stock, minStock, reorder, target, sfi, longDescription)
SP = [
 # ----- Caja reductora Twin Disc MG-5170DC -----
 ("CR-FIL-ACE-01","Filtro de aceite spin-on para caja Twin Disc MG-5170DC","Filtro","B","Twin Disc","MG-5170DC","XB3614A","ud",2,2,2,2,600,
  "Filtro de aceite tipo spin-on del circuito hidraulico de la caja. Reemplazo cada 1000 h o 6 meses. P/N Twin Disc XB3614A."),
 ("CR-STR-01","Colador/filtro de succion para caja Twin Disc MG-5170DC","Filtro","B","Twin Disc","MG-5170DC",None,"ud",1,1,1,1,600,
  "Colador de succion ubicado bajo la bomba de aceite. Limpiar en cada cambio de aceite; reemplazable si esta danado."),
 ("CR-ANODO-01","Anodos de zinc del intercambiador de calor para caja Twin Disc MG-5170DC","Anodo","B","Twin Disc","MG-5170DC",None,"jgo",2,2,2,4,600,
  "Anodos/varillas de zinc del intercambiador de calor. Controlar cada 90 dias; reemplazar si se consumio mas del 50%."),
 ("CR-ACE-01","Aceite de transmision para caja Twin Disc MG-5170DC","Lubricante","A","Twin Disc","MG-5170DC",None,"l",30,28,28,56,600,
  "Aceite de transmision segun placa de datos de la caja. Capacidad del carter 28 L mas mangueras e intercambiador."),
 ("CR-JUN-01","Kit de juntas y O-rings de servicio para caja Twin Disc MG-5170DC","Kit","B","Twin Disc","MG-5170DC",None,"kit",1,1,1,2,600,
  "Kit de juntas/O-rings de servicio (tapon O-ring de drenaje, breather/respiradero y tapas) para cambio de aceite e intervenciones menores."),
 ("CR-COUP-01","Elemento de acople torsional flexible para caja Twin Disc MG-5170DC","Acople","A","Twin Disc","MG-5170DC","VKE4911S / CF-R","ud",0,0,0,1,600,
  "Elemento del acople torsional flexible (Vulkan VKE4911S o Centa CF-R, 14\"/18\"). Repuesto critico estrategico; confirmar modelo por analisis torsional (TVA) del sistema."),
 ("CR-GRASA-01","Grasa de litio NLGI N2 para sellos de eje de salida (caja Twin Disc)","Lubricante","C","Twin Disc","MG-5170DC",None,"kg",1,1,1,2,600,
  "Grasa de litio NLGI N2 tipo bomba de agua para engrase de los sellos del eje de salida (cada 100 h)."),
 # ----- Generadores Cummins serie ES/C -----
 ("GEN-FIL-ACE-01","Filtro de aceite para generador Cummins serie ES/C","Filtro","B","Cummins","Serie ES/C (ES17D5/C17D5)",None,"ud",2,2,2,2,600,
  "Filtro de aceite del generador. Cambio en Revision B (cada 225-250 h / 3 meses). Confirmar P/N por modelo y N de serie del equipo."),
 ("GEN-FIL-COMB-01","Filtro de combustible para generador Cummins serie ES/C","Filtro","B","Cummins","Serie ES/C (ES17D5/C17D5)",None,"ud",2,2,2,2,600,
  "Filtro de combustible. Cambio en Revision B (cada 225-250 h / 3 meses). Drenar tambien el separador de agua/combustible."),
 ("GEN-FIL-AIRE-01","Elemento de filtro de aire para generador Cummins serie ES/C","Filtro","B","Cummins","Serie ES/C (ES17D5/C17D5)",None,"ud",2,2,2,2,600,
  "Elemento de filtro de aire. Reemplazar segun indicador de vacuometro / Revision C. Revisar carcasa, mangueras y abrazaderas."),
 ("GEN-COR-VENT-01","Correa del ventilador para generador Cummins serie ES/C","Correa","B","Cummins","Serie ES/C (ES17D5/C17D5)",None,"ud",1,1,1,2,600,
  "Correa del ventilador. Revisar tension en Revision C; cambio en Revision D (cada 1450-1500 h / 12 meses)."),
 ("GEN-ACE-01","Aceite lubricante CH4 15W40 para generador Cummins serie ES/C","Lubricante","A","Cummins","Serie ES/C (ES17D5/C17D5)",None,"l",20,15,15,40,600,
  "Aceite lubricante CH4 15W40 (no usar aceite de motores a gasolina en motores diesel). Para reposicion y cambios de aceite."),
 ("GEN-REF-01","Liquido refrigerante premezclado para generador Cummins serie ES/C","Refrigerante","B","Cummins","Serie ES/C (ES17D5/C17D5)",None,"l",10,5,5,20,600,
  "Liquido refrigerante premezclado para rellenar el radiador y el reservorio. Cambio de refrigerante en Revision D."),
 ("GEN-MANG-COMB-01","Mangueras de combustible para generador Cummins serie ES/C","Manguera","B","Cummins","Serie ES/C (ES17D5/C17D5)",None,"jgo",1,1,1,1,600,
  "Juego de mangueras de combustible. Reemplazar segun condicion / Revision D. Revisar conexiones y juntas de las lineas."),
 ("GEN-INY-01","Inyector / kit de inyeccion para generador Cummins serie ES/C","Inyector","A","Cummins","Serie ES/C (ES17D5/C17D5)",None,"ud",0,0,0,1,600,
  "Inyectores / kit de inyeccion. Calibracion de inyectores y bomba inyectora en Revision E (Servicio Tecnico Bosch). Repuesto critico; confirmar P/N por N de serie."),
 ("GEN-BAT-01","Bateria de arranque para generador Cummins serie ES/C","Bateria","A","Cummins","Serie ES/C (ES17D5/C17D5)",None,"ud",1,1,1,2,600,
  "Bateria de arranque. Controlar terminales, densidad de electrolito y voltaje de cargador (14 V en carga, 13,2 V en reposo)."),
 ("GEN-POLEA-VENT-01","Polea del ventilador para generador Cummins serie ES/C","Repuesto","C","Cummins","Serie ES/C (ES17D5/C17D5)",None,"ud",0,0,0,1,600,
  "Polea del ventilador. Revisar/cambiar segun Revision D-E. Verificar aspas del ventilador."),
]

# columnas destino en la hoja spares
def setsp(r, sku, name, cat, crit, mfg, model, pn, unit, stock, mn, ro, tg, sfi, longd):
    m = {"A":"M01","B":sku,"C":name,"D":cat,"E":crit,"F":"ACTIVE","G":mfg,"H":model,
         "J":pn,"K":unit,"L":stock,"M":mn,"N":ro,"O":tg,"P":"Panol Maquinas","Q":sfi,
         "S":longd,"Y":False,"AB":TENANT}
    for L, v in m.items():
        ws2.cell(row=r, column=C(L)).value = v

r = ws2.max_row + 1
addsp = 0
for s in SP:
    setsp(r, *s)
    r += 1; addsp += 1
io.write(f"repuestos nuevos agregados: {addsp}\n".encode("utf-8"))
wb2.save(REP)
io.write(b"REPUESTOS guardado.\n")
io.write(b"--OK--\n")
