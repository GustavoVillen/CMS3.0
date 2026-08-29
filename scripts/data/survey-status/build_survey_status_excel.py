# -*- coding: utf-8 -*-
"""Genera el Excel de Survey Status de la flota a partir de los PDF de clase."""
import os, re, glob, datetime
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import survey_status_parser as B

FOLDER = B.FOLDER
OUT = os.path.join(FOLDER, "Survey Status Flota.xlsx")
HOY = datetime.date(2026, 8, 29)

# ---- recolectar ----
files = sorted(glob.glob(os.path.join(FOLDER, "*hip *tatus*.pdf")))
seen = set()
vessels = []
for f in files:
    base = os.path.basename(f)
    code = base.split("-")[0].upper()
    if code in seen:
        continue
    with pdfplumber.open(f) as pdf:
        head = pdf.pages[0].extract_text() or ""
        full = "\n".join((p.extract_text() or "") for p in pdf.pages)
    v = B.parse_nk(f) if "NIPPON KAIJI" in head else B.parse_rina(f)
    if v["soc"] == "ClassNK" and v["name"]:
        v["name"] = re.sub(r"(\S)\1", lambda m: m.group(1), v["name"]).strip()
    if not v["name"]:
        v["name"] = code
    seen.add(code)
    row = B.build_row(code, v)
    row["_file"] = base
    row["_v"] = v
    if "renovacion" in base.lower():
        row["notes"].insert(0, "Certificado de clase VENCIDO - en renovacion de clase a seco")
    # RINA: ordinaria / anual
    if v["soc"] == "RINA":
        S = [r for r in v["surveys"] if r.get("section") == "CLASS SURVEYS"]
        ho = [r for r in S if re.match(r"^Hull Ordinary", r["type"], re.I)]
        row["an_last"] = B.d_rina(ho[0]["last"]) if ho and ho[0]["last"] else None
        row["an_due"] = B.d_rina(ho[0]["due"]) if ho and ho[0]["due"] else None
        row["an_lim"] = (B.range_end_rina(ho[0]["range"]) if ho else None) or row["an_due"]
    else:
        an = [r for r in v["surveys"] if re.match(r"^Annual Survey", r["kind"], re.I)]
        def pk(r, k):
            val = r[k].replace("--", "").strip()
            return B.d_nk(val) if val else None
        row["an_last"] = pk(an[0], "last") if an else None
        row["an_due"] = pk(an[0], "due") if an else None
        m = re.search(r"Last Partial Survey Date(\d{1,2} \w{3} \d{4})", full)
        if m:
            row["notes"].append("Ultima inspeccion parcial de ejes: " +
                                B.fmt(B.d_nk(m.group(1))))
    vessels.append(row)

# normalizacion de codigo y tipo
TIPOS = {"tug": "Remolcador", "Pusher": "Empujador", "barge - oil": "Barcaza tanque"}
INFERIDOS = {"LTE": "Empujador", "MGT10": "Barcaza", "MGT11": "Barcaza", "MGT12": "Barcaza",
             "MGT13": "Barcaza", "MGT14": "Barcaza", "MGT15": "Barcaza"}
DATE_RE = re.compile(r"(\d{2})(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(\d{4})")

def lindo(txt):
    txt = DATE_RE.sub(lambda m: "%s/%02d/%s" % (m.group(1), B.MON[m.group(2)], m.group(3)), txt)
    return re.sub(r"(\d{1,2}) (Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) (\d{4})",
                  lambda m: "%02d/%02d/%s" % (int(m.group(1)), B.MON[m.group(2)], m.group(3)), txt)

for v in vessels:
    if v["code"] == "DHC":
        v["code"] = "DCH"
    t = TIPOS.get(v["type"], v["type"])
    if not t:
        t = INFERIDOS.get(v["code"], "")
    v["type"] = t
    v["notes"] = [lindo(n) for n in v["notes"]]

# orden: primero remolcadores/empujadores, luego barcazas por numero
def sort_key(r):
    m = re.match(r"MGT(\d+)", r["code"])
    return (1, int(m.group(1))) if m else (0, r["code"])
vessels.sort(key=sort_key)

# ---- estilos ----
AZUL = "1F3864"; AZUL2 = "2E5395"; GRIS = "F2F2F2"; BLANCO = "FFFFFF"
ROJO = "FFC7CE"; ROJO_T = "9C0006"
NARANJA = "FFE0B2"; NARANJA_T = "8A4B00"
AMARILLO = "FFF2CC"; AMARILLO_T = "7F6000"
VERDE = "E2EFDA"; VERDE_T = "375623"

thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "Survey Status"

ws["A1"] = "SURVEY STATUS DE CLASE - FLOTA"
ws["A1"].font = Font(bold=True, size=15, color=AZUL)
ws["A2"] = ("Fuente: Ship Status emitidos por RINA (20/08/2026) y ClassNK (20-21/08/2026). "
            "Fechas en formato dd/mm/aaaa. Corte de analisis: 29/08/2026.")
ws["A2"].font = Font(size=9, italic=True, color="595959")

GROUPS = [
    ("", ["Buque", "Codigo", "Tipo", "Sociedad de clasificacion"]),
    ("RENOVACION / SPECIAL SURVEY", ["Ultima", "Vence"]),
    ("INTERMEDIA", ["Ultima", "Vence"]),
    ("PERIODICA / ANUAL", ["Ultima", "Vence"]),
    ("INSPECCION EN SECO (VARADA)", ["Ultima", "Vence"]),
    ("EJE PORTAHELICE", ["Ultima", "Vence"]),
    ("", ["Observaciones"]),
]
R1, R2 = 4, 5
col = 1
for gname, subs in GROUPS:
    start = col
    for s in subs:
        c = ws.cell(row=R2, column=col, value=s)
        c.font = Font(bold=True, color=BLANCO, size=10)
        c.fill = PatternFill("solid", fgColor=AZUL2)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = box
        col += 1
    end = col - 1
    if gname:
        ws.merge_cells(start_row=R1, start_column=start, end_row=R1, end_column=end)
    cc = ws.cell(row=R1, column=start, value=gname)
    cc.font = Font(bold=True, color=BLANCO, size=10)
    cc.fill = PatternFill("solid", fgColor=AZUL)
    cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for x in range(start, end + 1):
        ws.cell(row=R1, column=x).fill = PatternFill("solid", fgColor=AZUL)
        ws.cell(row=R1, column=x).border = box
ws.row_dimensions[R1].height = 30
ws.row_dimensions[R2].height = 24


def estado(due):
    if not due:
        return None
    d = (due - HOY).days
    if d < 0:
        return (ROJO, ROJO_T)
    if d <= 90:
        return (NARANJA, NARANJA_T)
    if d <= 180:
        return (AMARILLO, AMARILLO_T)
    return (VERDE, VERDE_T)


r = R2 + 1
for v in vessels:
    vals = [v["name"], v["code"], v["type"] or "", v["soc"],
            v["ss_last"], v["ss_due"], v["im_last"], v["im_due"],
            v.get("an_last"), v.get("an_due"), v["dd_last"], v["dd_due"],
            v["ts_last"], v["ts_due"], " | ".join(v["notes"])]
    for i, val in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=val if val not in (None, "") else None)
        c.border = box
        c.font = Font(size=10)
        if isinstance(val, datetime.date):
            c.number_format = "DD/MM/YYYY"
            c.alignment = Alignment(horizontal="center")
        if i == 1:
            c.font = Font(size=10, bold=True)
        if i in (2, 3, 4):
            c.alignment = Alignment(horizontal="center")
        if i == 15:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=8, color="595959")
        if val is None or val == "":
            if i in (5, 6, 7, 8, 9, 10, 11, 12, 13, 14):
                c.value = "-"
                c.alignment = Alignment(horizontal="center")
                c.font = Font(size=10, color="A6A6A6")
    # semaforo: se pinta segun la ultima fecha admisible (fin de ventana)
    LIM = {6: "ss_lim", 8: "im_lim", 10: "an_lim", 12: "dd_lim", 14: "ts_lim"}
    DUE = {6: "ss_due", 8: "im_due", 10: "an_due", 12: "dd_due", 14: "ts_due"}
    for cidx in (6, 8, 10, 12, 14):
        lim = v.get(LIM[cidx]) or v.get(DUE[cidx])
        if isinstance(lim, datetime.date) and isinstance(ws.cell(row=r, column=cidx).value, datetime.date):
            est = estado(lim)
            if est:
                ws.cell(row=r, column=cidx).fill = PatternFill("solid", fgColor=est[0])
                ws.cell(row=r, column=cidx).font = Font(size=10, bold=True, color=est[1])
    if (r - R2) % 2 == 0:
        for i in range(1, 16):
            cc = ws.cell(row=r, column=i)
            if cc.fill.fgColor.rgb in (None, "00000000"):
                cc.fill = PatternFill("solid", fgColor=GRIS)
    ws.row_dimensions[r].height = 30
    r += 1

last_row = r - 1
widths = [20, 9, 14, 13, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 62]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "E6"
ws.auto_filter.ref = "A%d:O%d" % (R2, last_row)

# leyenda
lr = last_row + 2
ws.cell(row=lr, column=1, value="REFERENCIAS").font = Font(bold=True, size=10, color=AZUL)
leg = [
    ("Vencido", ROJO, ROJO_T),
    ("Vence dentro de 90 dias", NARANJA, NARANJA_T),
    ("Vence entre 90 y 180 dias", AMARILLO, AMARILLO_T),
    ("Vence en mas de 180 dias", VERDE, VERDE_T),
]
for i, (txt, bg, fg) in enumerate(leg):
    c = ws.cell(row=lr + 1 + i, column=1, value=txt)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(size=9, bold=True, color=fg)
    c.border = box
notas = [
    "Renovacion / Special Survey: renovacion de clase (casco y maquinas). En RINA se toma la fila 'Hull Renewal'; si casco y maquinas difieren se aclara en Observaciones.",
    "Intermedia: 'Hull Intermediate' (RINA) / 'Intermediate Survey' (ClassNK).",
    "Periodica / Anual: 'Hull Ordinary' (RINA) / 'Annual Survey' (ClassNK).",
    "Inspeccion en seco: 'Bottom Dry Condition' (RINA) / 'Docking Survey' (ClassNK).",
    "Eje portahelice: 'Tailshaft - Propeller shaft' (RINA) / 'Prop. Shaft Svy - Ordinary Svy.' (ClassNK). Las barcazas no tienen propulsion propia: figura '-'.",
    "El color de las columnas 'Vence' se calcula con la ultima fecha admisible: si la sociedad de clasificacion concede una ventana, se toma el cierre de esa ventana (ver Observaciones).",
    "Una celda con '-' significa que el certificado no registra esa fecha (por ejemplo, una inspeccion que aun no se hizo en el ciclo de clase vigente).",
    "Los certificados de ClassNK no indican notacion de servicio: el tipo de LATERE y de MGT 10 a MGT 15 se deduce del propio certificado (tiene o no inspeccion de ejes).",
]
for i, n in enumerate(notas):
    c = ws.cell(row=lr + 1 + len(leg) + i, column=1, value=n)
    c.font = Font(size=8, italic=True, color="595959")

# ---- hoja de detalle ----
ws2 = wb.create_sheet("Detalle por buque")
ws2["A1"] = "DETALLE DE INSPECCIONES SEGUN CERTIFICADO"
ws2["A1"].font = Font(bold=True, size=13, color=AZUL)
hdr = ["Buque", "Codigo", "Sociedad", "Inspeccion", "Observacion del certificado",
       "Ultima", "Vence", "Ventana / Rango", "Estado"]
for i, h in enumerate(hdr, start=1):
    c = ws2.cell(row=3, column=i, value=h)
    c.font = Font(bold=True, color=BLANCO, size=10)
    c.fill = PatternFill("solid", fgColor=AZUL2)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
rr = 4
for v in vessels:
    src = v["_v"]
    if src["soc"] == "RINA":
        for s in src["surveys"]:
            if s.get("section") != "CLASS SURVEYS":
                continue
            vals = [v["name"], v["code"], "RINA", s["type"], s["remarks"],
                    B.d_rina(s["last"]) if s["last"] else None,
                    B.d_rina(s["due"]) if s["due"] else None,
                    s["range"] if s["range"].strip() not in ("", "-") else "",
                    s["status"]]
            for i, val in enumerate(vals, start=1):
                c = ws2.cell(row=rr, column=i, value=val if val not in (None, "") else None)
                c.border = box
                c.font = Font(size=9)
                if isinstance(val, datetime.date):
                    c.number_format = "DD/MM/YYYY"
                    c.alignment = Alignment(horizontal="center")
            rr += 1
    else:
        for s in src["surveys"]:
            kind = (s.get("group", "") + " / " + s["kind"]).strip(" /")
            if not re.search(r"Survey|Svy", kind):
                continue
            def pk(k):
                val = s[k].replace("--", "").strip()
                return B.d_nk(val) if val else None
            vals = [v["name"], v["code"], "ClassNK", kind, "",
                    pk("last"), pk("due"),
                    s["range"].replace("--", "").strip(), s["status"].replace("--", "").strip()]
            for i, val in enumerate(vals, start=1):
                c = ws2.cell(row=rr, column=i, value=val if val not in (None, "") else None)
                c.border = box
                c.font = Font(size=9)
                if isinstance(val, datetime.date):
                    c.number_format = "DD/MM/YYYY"
                    c.alignment = Alignment(horizontal="center")
            rr += 1
for i, w in enumerate([20, 9, 11, 42, 30, 12, 12, 26, 14], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A4"
ws2.auto_filter.ref = "A3:I%d" % (rr - 1)

wb.save(OUT)
print("OK ->", OUT)
print("buques:", len(vessels), "| filas detalle:", rr - 4)
