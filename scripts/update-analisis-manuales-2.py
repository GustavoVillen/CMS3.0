# -*- coding: utf-8 -*-
"""
Segunda tanda de actualizacion de las planillas de analisis con manuales:
- Alternadores CRAMACO G2R (M01-ALT-BR / M01-ALT-ER)
- Electrocompresor a tornillo CETEC (M01-COMP-NK40; manual NK60-15/20)
- Purificadora de combustible Alfa Laval MAB 102B-14 (M01-PURIF) -> solo repuestos
Hace backup de ambos archivos antes de tocarlos.
"""
import openpyxl, shutil, datetime, sys

io = sys.stdout.buffer
MANT = r"C:\CMS3.0\MisDocs\MANTENIMIENTO EN ANALISIS.xlsx"
REP  = r"C:\CMS3.0\MisDocs\REPUESTOS ANALISIS.xlsx"
stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
for p in (MANT, REP):
    bak = p.replace(".xlsx", f".bak-{stamp}.xlsx")
    shutil.copyfile(p, bak)
    io.write(f"backup -> {bak}\n".encode("utf-8"))

def C(L): return openpyxl.utils.column_index_from_string(L)
TENANT = "cmqhafl85006ysll4f0hecqcu"

ASSET = {
 "M01-ALT-BR":    dict(sg=8, sf=800, aid="cmqhe9y33003628l4ba89i0ic", an="Alternador Babor",                  trm="AUTO_WO"),
 "M01-ALT-ER":    dict(sg=8, sf=800, aid="cmqhe9y2s003228l4i6yy7hr4", an="Alternador Estribor",               trm="AUTO_WO"),
 "M01-COMP-NK40": dict(sg=7, sf=700, aid="cmqhe9y7c005a28l4855yhpd7", an="Electrocompresor a tornillo NK40",  trm="AUTO_WO"),
}

# =========================================================
# 1) MANTENIMIENTO
# =========================================================
wb = openpyxl.load_workbook(MANT)
ws = wb["maintenance_plans"]

CORR = {
 ("M01-COMP-NK40","M01-COMP-NK40-05"): {
    "H":"Cambio de aceite sintetico y filtro separador (cada 3000 h)","J":"MAINTENANCE",
    "I":"CETEC NK - aceite sintetico cada 3000 h (primer cambio a las 500 h). El filtro separador se cambia junto con el aceite. Capacidad 6 L de CETEC Synthetic Oil."},
}
applied = 0
for r in range(2, ws.max_row+1):
    ac = ws.cell(row=r, column=C("B")).value
    tc = ws.cell(row=r, column=C("G")).value
    if (ac, tc) in CORR:
        for L, v in CORR[(ac, tc)].items():
            ws.cell(row=r, column=C(L)).value = v
        applied += 1
io.write(f"correcciones aplicadas: {applied}/{len(CORR)}\n".encode("utf-8"))

# (assetCode, suffix, title, taskType, triggerType, months, hours, estHours, desc)
NEW = [
 # ---- Alternadores CRAMACO G2R ----
 *[(a, 4, "Cambio de rodamientos sellados","MAINTENANCE","HOURS",None,20000,4,
    "CRAMACO G2R - rodamientos sellados prelubricados, libres de mantenimiento; reemplazar a las 20000 h o antes ante altas temperaturas o sobrevelocidad.") for a in ("M01-ALT-BR","M01-ALT-ER")],
 *[(a, 5, "Verificacion de diodos rotativos, rectificadores y regulador de tension (AVR)","INSPECTION","MONTHS",12,None,1,
    "CRAMACO G2R - control del puente de diodos rotativos, rectificadores y AVR ('VOLTS'); verificar conexiones y estado segun la seccion de fallas del manual.") for a in ("M01-ALT-BR","M01-ALT-ER")],
 # ---- Electrocompresor CETEC NK ----
 ("M01-COMP-NK40", 14, "Verificacion de nivel de aceite","INSPECTION","HOURS",None,100,1,
    "CETEC NK - verificar nivel antes de cada arranque y cada 100 h. Capacidad 6 L."),
 ("M01-COMP-NK40", 15, "Limpieza del enfriador aire/aceite","MAINTENANCE","HOURS",None,300,1,
    "CETEC NK - limpiar las aletas del enfriador con aire comprimido cada 300 h."),
 ("M01-COMP-NK40", 16, "Limpieza a seco del filtro de aire","INSPECTION","HOURS",None,500,1,
    "CETEC NK - limpieza a seco cada 500 h; observar el indicador semanalmente; cambiar al menos 1 vez al ano; no reutilizar tras 3 limpiezas."),
 ("M01-COMP-NK40", 17, "Cambio del filtro de aceite","MAINTENANCE","HOURS",None,1500,1,
    "CETEC NK - cambio del filtro de aceite (unidad sellada) cada 1500 h."),
 ("M01-COMP-NK40", 18, "Cambio del filtro de aire","MAINTENANCE","HOURS",None,1500,1,
    "CETEC NK - cambio del cartucho de filtro de aire cada 1500 h."),
 ("M01-COMP-NK40", 19, "Control y ajuste de correas (transmision Poly-V)","INSPECTION","MONTHS",12,None,1,
    "CETEC NK - observacion anual del estado de las correas; flecha maxima 5 mm; verificar alineacion de poleas."),
 ("M01-COMP-NK40", 20, "Prueba de la valvula de seguridad","INSPECTION","MONTHS",12,None,1,
    "CETEC NK - probar tras cada mantenimiento; verificar descarga de presion y ausencia de fugas."),
]

row = ws.max_row + 1
added = 0
for (ac, suf, title, ttype, trig, months, hours, est, desc) in NEW:
    a = ASSET[ac]
    tc = f"{ac}-{suf:02d}"
    vals = {"A":"M01","B":ac,"C":a["sg"],"D":a["sf"],"E":a["aid"],"F":a["an"],
            "G":tc,"H":title,"I":desc,"J":ttype,"K":trig,"L":months,"M":hours,
            "N":est,"O":"Jefe de Maquinas","V":a["trm"],"W":"AUTO","AG":"ACTIVE","AT":TENANT}
    for L, v in vals.items():
        ws.cell(row=row, column=C(L)).value = v
    row += 1; added += 1
io.write(f"tareas nuevas agregadas: {added}\n".encode("utf-8"))
wb.save(MANT)
io.write(b"MANTENIMIENTO guardado.\n")

# =========================================================
# 2) REPUESTOS
# =========================================================
wb2 = openpyxl.load_workbook(REP)
ws2 = wb2["spares"]

# (sku, name, cat, crit, mfg, model, pn, unit, stock, min, reorder, target, sfi, longDesc)
SP = [
 # ----- Alternadores CRAMACO G2R (sin P/N en el manual) -----
 ("ALT-DIODO-01","Puente de diodos rotativos / rectificador rotante para alternador CRAMACO G2R","Electrico","A","CRAMACO","G2R 160/200/250/315/400",None,"ud",1,1,1,2,800,
  "Puente de diodos rotativos del alternador brushless. Repuesto critico; su falla provoca perdida de excitacion y de tension. Confirmar P/N por N de serie del alternador."),
 ("ALT-AVR-01","Regulador automatico de tension (AVR tipo 'VOLTS') para alternador CRAMACO G2R","Electronico","A","CRAMACO","G2R 160/200/250/315/400",None,"ud",1,1,1,1,800,
  "Regulador de tension (AVR) identificado como 'VOLTS'. Repuesto critico; su falla afecta la regulacion y estabilidad de tension. Confirmar P/N por N de serie."),
 ("ALT-ROD-01","Juego de rodamientos sellados para alternador CRAMACO G2R","Rodamiento","B","CRAMACO","G2R 160/200/250/315/400",None,"jgo",1,1,1,2,800,
  "Rodamientos sellados prelubricados, libres de mantenimiento. Reemplazo a las 20000 h o antes ante altas temperaturas/sobrevelocidad."),
 ("ALT-VARIS-01","Varistor de proteccion para alternador CRAMACO G2R","Electrico","C","CRAMACO","G2R 160/200/250/315/400",None,"ud",1,1,1,2,800,
  "Varistor de proteccion del sistema de excitacion. Repuesto economico de proteccion frente a sobretensiones."),
 # ----- Electrocompresor CETEC NK (codigos del manual NK60-15/20; el activo es NK40) -----
 ("COMP-FIL-AIRE-01","Filtro de aire (elemento) para electrocompresor CETEC NK","Filtro","B","CETEC","NK40 (datos manual NK60-15/20)","13053","ud",2,2,2,2,700,
  "Cartucho de filtro de aire. Cambio cada 1500 h; limpieza a seco cada 500 h. Codigo CETEC 13053 segun manual NK60-15/20; confirmar equivalencia para NK40 por N de serie."),
 ("COMP-FIL-ACE-01","Filtro de aceite (unidad sellada) para electrocompresor CETEC NK","Filtro","B","CETEC","NK40 (datos manual NK60-15/20)","199018","ud",2,2,2,2,700,
  "Filtro de aceite tipo unidad sellada. Cambio cada 1500 h. Codigo CETEC 199018 segun manual NK60-15/20; confirmar equivalencia para NK40."),
 ("COMP-FIL-SEP-01","Filtro separador (elemento filtrante) para electrocompresor CETEC NK","Filtro","B","CETEC","NK40 (datos manual NK60-15/20)","240154","ud",1,1,1,2,700,
  "Elemento del filtro separador aire/aceite. Se cambia junto con el aceite (cada 3000 h). Codigo CETEC 240154 segun manual NK60-15/20; confirmar para NK40."),
 ("COMP-VALV-SEG-01","Valvula de seguridad (ajustable 4-16 bar) para electrocompresor CETEC NK","Valvula","B","CETEC","NK40 (datos manual NK60-15/20)","240058","ud",1,1,1,1,700,
  "Valvula de seguridad ajustable de 4 a 16 bar. Probar tras cada mantenimiento. Codigo CETEC 240058 segun manual NK60-15/20."),
 ("COMP-ACE-01","Aceite CETEC Synthetic Oil (x 5 L) para electrocompresor CETEC NK","Lubricante","A","CETEC","NK40 (datos manual NK60-15/20)","208002","l",10,6,6,15,700,
  "Aceite sintetico CETEC Synthetic Oil. Cambio cada 3000 h (primer cambio a 500 h). Capacidad del deposito 6 L. No mezclar especificaciones ni fabricantes."),
 ("COMP-CORREA-01","Correa Poly-V seccion J (50 Hz) para electrocompresor CETEC NK","Correa","B","CETEC","NK40 (datos manual NK60-15/20)","162574-12","ud",1,1,1,2,700,
  "Correa de transmision Poly-V seccion J. Control anual; cambio por desgaste. Codigo CETEC 162574-12 (NK60-15, 8 bar) segun manual; confirmar para NK40."),
 ("COMP-SENS-P-01","Sensor de presion para electrocompresor CETEC NK","Sensor","C","CETEC","NK40 (datos manual NK60-15/20)","339012","ud",1,1,1,1,700,
  "Sensor de presion de la central electronica de comando. Codigo CETEC 339012 segun manual NK60-15/20."),
 ("COMP-SENS-T-01","Sensor de temperatura para electrocompresor CETEC NK","Sensor","C","CETEC","NK40 (datos manual NK60-15/20)","339013","ud",1,1,1,1,700,
  "Sensor de temperatura de la central electronica de comando. Codigo CETEC 339013 segun manual NK60-15/20."),
 ("COMP-CENT-01","Central electronica de comando (modelo P1) para electrocompresor CETEC NK","Electronico","B","CETEC","NK40 (datos manual NK60-15/20)","339038","ud",0,0,0,1,700,
  "Central electronica de comando y control modelo P1. Repuesto critico estrategico. Codigo CETEC 339038 segun manual NK60-15/20."),
 # ----- Purificadora Alfa Laval MAB 102B-14 -----
 ("PUR-KIT-INT-01","Intermediate service kit para purificadora Alfa Laval MAB 102B-14","Kit","A","Alfa Laval","MAB 102B-14","549301-01","kit",1,1,1,1,700,
  "Kit de servicio intermedio (O-rings, juntas, friction pads, fixing rings). Para el servicio intermedio del separador. P/N Alfa Laval 549301-01."),
 ("PUR-KIT-MAJ-01","Major service kit para purificadora Alfa Laval MAB 102B-14","Kit","A","Alfa Laval","MAB 102B-14","549302-01","kit",1,1,1,1,700,
  "Kit de servicio mayor (rodamiento, bearing pin, bushings, juntas). Para el servicio mayor del separador. P/N Alfa Laval 549302-01."),
 ("PUR-IMP-01","Impeller de bomba de alimentacion/descarga para purificadora Alfa Laval MAB 102B-14","Impeller","B","Alfa Laval","MAB 102B-14","547995-01","ud",1,1,1,2,700,
  "Impeller de la bomba de alimentacion/descarga del separador. P/N Alfa Laval 547995-01 (alternativo 547996-01)."),
 ("PUR-COUP-01","Shear-pin coupling para purificadora Alfa Laval MAB 102B-14","Acople","B","Alfa Laval","MAB 102B-14","535118-01","ud",1,1,1,2,700,
  "Acople con pasador de corte (shear-pin coupling) de la bomba. Elemento de seguridad mecanica. P/N Alfa Laval 535118-01."),
]

def setsp(r, sku, name, cat, crit, mfg, model, pn, unit, stock, mn, ro, tg, sfi, longd):
    m = {"A":"M01","B":sku,"C":name,"D":cat,"E":crit,"F":"ACTIVE","G":mfg,"H":model,
         "J":pn,"K":unit,"L":stock,"M":mn,"N":ro,"O":tg,"P":"Panol Maquinas","Q":sfi,
         "S":longd,"Y":False,"AB":TENANT}
    for L, v in m.items():
        ws2.cell(row=r, column=C(L)).value = v

r = ws2.max_row + 1
addsp = 0
for s in SP:
    setsp(r, *s)
    r += 1; addsp += 1
io.write(f"repuestos nuevos agregados: {addsp}\n".encode("utf-8"))
wb2.save(REP)
io.write(b"REPUESTOS guardado.\n--OK--\n")
