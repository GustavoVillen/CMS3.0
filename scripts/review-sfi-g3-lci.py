# Genera la planilla de revisión para reubicar los equipos y planes de LCI y
# Salvamento en el grupo SFI 3.
#
# CRITERIO: no se inventa una taxonomía nueva. El DON CHICUETO (DCH) ya tiene su
# grupo 3 armado exactamente como "LCI y Salvamento" (balsas, bombas de incendio,
# extintores, CO2, detección, cortes de emergencia, motor de lancha). Se toma ese
# grupo como modelo y se propone alinear al resto de la flota POR NOMBRE DE
# EQUIPO: si un equipo se llama igual que uno que en DCH está en G3, se propone
# G3 también en los demás buques.
#
# Lo que hoy ocupa G3 sin pertenecer al modelo (equipos de izaje) se marca para
# revisión aparte, con G2 "Sistemas de Carga" como sugerencia.
#
# Los planes siguen al equipo: se propone para el plan el grupo propuesto de su
# equipo. NO escribe nada: sólo genera el Excel.

import csv, io, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

GRUPOS = {
    "0": "Inspecciones y Pruebas", "1": "Casco y Estructuras", "2": "Sistemas de Carga",
    "3": "LCI y Salvamento", "4": "Sistemas de Navegación", "5": "Sistemas de Habitabilidad",
    "6": "Sistemas de Propulsión y Generación", "7": "Sistemas Auxiliares",
    "8": "Sistemas Eléctricos", "9": "Sistemas de Automatización y Control",
}

# Equipos de izaje que hoy ocupan G3 y no son LCI ni salvamento.
IZAJE = ("elevaci", "malacate", "monorriel", "aparejo", "eslinga")

# El modelo de DCH por sí solo no alcanza: hay equipos de LCI que en DCH no
# existen con ese nombre exacto (p. ej. "Sistema de Lucha Contra Incendio", que
# son 33 equipos en el resto de la flota). Se suma un cruce por palabra clave.
LCI = ("incendio", "salvamento", "salvavida", "extintor", "co2", "balsa",
       "lci", "rescate", "contra incendio")

assets = list(csv.DictReader(io.open("scratch-assets.csv", encoding="utf-8")))
plans  = list(csv.DictReader(io.open("scratch-plans.csv",  encoding="utf-8")))

def norm(s): return " ".join((s or "").strip().lower().split())

# El modelo: nombres que en DCH están en G3.
modelo = {norm(a["name"]) for a in assets if a["vesselCode"] == "DCH" and a["g"] == "3"}

def propuesta(nombre, g):
    n = norm(nombre)
    if n in modelo or any(k in n for k in LCI):
        return "3"
    if g == "3":
        return "2" if any(k in n for k in IZAJE) else ""
    return ""

for a in assets:
    a["prop"] = propuesta(a["name"], a["g"])
prop_by_asset = {a["id"]: a["prop"] for a in assets}
for p in plans:
    p["prop"] = prop_by_asset.get(p["assetId"], "")

a_ch = [a for a in assets if a["prop"] and a["prop"] != a["g"]]
p_ch = [p for p in plans  if p["prop"] and p["prop"] != p["g"]]

wb = Workbook()
HDR  = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F4E79")
WARN = PatternFill("solid", fgColor="FFF2CC")

def sheet(ws, headers, rows, widths):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = HDR
        ws.cell(1, c).fill = FILL
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

ws = wb.active
ws.title = "RESUMEN"
resumen = [
    ["Planilla de revisión — reubicación a G3 'LCI y Salvamento'", ""],
    ["", ""],
    ["Criterio", "Alinear la flota con el DON CHICUETO, que ya tiene el G3 bien armado."],
    ["", "Se compara POR NOMBRE de equipo contra los que en DCH están en G3."],
    ["", ""],
    ["Equipos a reubicar", len(a_ch)],
    ["Planes a reubicar", len(p_ch)],
    ["", ""],
    ["Movimiento", "Equipos"],
]
mov_a = collections.Counter("G%s -> G%s" % (a["g"] or "?", a["prop"]) for a in a_ch)
for k, v in sorted(mov_a.items()):
    resumen.append([k, v])
resumen += [["", ""], ["Movimiento", "Planes"]]
mov_p = collections.Counter("G%s -> G%s" % (p["g"] or "?", p["prop"]) for p in p_ch)
for k, v in sorted(mov_p.items()):
    resumen.append([k, v])
resumen += [
    ["", ""],
    ["Cómo revisar", "Poné NO en la columna ¿APLICAR? de las filas que no quieras mover."],
    ["", "Si querés otro grupo, escribí el número en GRUPO PROPUESTO."],
]
for row in resumen:
    ws.append(row)
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 78
ws["A1"].font = Font(bold=True, size=13)

sheet(wb.create_sheet("EQUIPOS"),
      ["BUQUE", "CÓDIGO", "EQUIPO", "GRUPO ACTUAL", "NOMBRE ACTUAL", "GRUPO PROPUESTO", "NOMBRE PROPUESTO", "¿APLICAR?"],
      [[a["vesselCode"], a["assetCode"], a["name"], "G" + (a["g"] or "?"), GRUPOS.get(a["g"], "—"),
        "G" + a["prop"], GRUPOS[a["prop"]], "SI"] for a in a_ch],
      [10, 22, 46, 14, 30, 17, 30, 11])

sheet(wb.create_sheet("PLANES"),
      ["BUQUE", "TASK ID", "EQUIPO", "TAREA", "GRUPO ACTUAL", "NOMBRE ACTUAL", "GRUPO PROPUESTO", "NOMBRE PROPUESTO", "¿APLICAR?"],
      [[p["vesselCode"], p["taskCode"], p["asset_name"], p["title"], "G" + (p["g"] or "?"), GRUPOS.get(p["g"], "—"),
        "G" + p["prop"], GRUPOS[p["prop"]], "SI"] for p in p_ch],
      [10, 20, 40, 52, 14, 28, 17, 28, 11])

# Los que hoy están en G3 y no encajan en el modelo: decisión aparte.
raros = [a for a in assets if a["g"] == "3" and norm(a["name"]) not in modelo]
sheet(wb.create_sheet("REVISAR G3 ACTUAL"),
      ["BUQUE", "CÓDIGO", "EQUIPO", "GRUPO ACTUAL", "SUGERENCIA", "COMENTARIO"],
      [[a["vesselCode"], a["assetCode"], a["name"], "G3",
        ("G" + a["prop"] + " " + GRUPOS[a["prop"]]) if a["prop"] else "—",
        "Está en LCI y Salvamento pero es equipo de izaje"] for a in raros],
      [10, 22, 46, 14, 34, 52])
for r in range(2, len(raros) + 2):
    for c in range(1, 7):
        wb["REVISAR G3 ACTUAL"].cell(r, c).fill = WARN

wb.save("Revision_SFI_G3_LCI.xlsx")

print("equipos a mover:", len(a_ch))
print("planes a mover :", len(p_ch))
print("en G3 a revisar:", len(raros))
print()
print("Modelo DCH G3 (%d equipos):" % len(modelo))
for n in sorted(modelo):
    print("  -", n)
print()
for k, v in sorted(mov_a.items()):
    print("ASSET %-12s %d" % (k, v))
for k, v in sorted(mov_p.items()):
    print("PLAN  %-12s %d" % (k, v))
