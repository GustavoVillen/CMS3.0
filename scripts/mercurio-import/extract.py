#!/usr/bin/env python3
"""
Extract data from MANTENIMIENTO Excel files into structured JSON.

Output: scripts/mercurio-import/data/*.json
"""
import json
import os
import sys
from pathlib import Path

import openpyxl

ROOT = Path(r"c:\NPMS\GPMS\MANTENIMIENTO")
OUT  = Path(r"c:\NPMS\GPMS\scripts\mercurio-import\data")
OUT.mkdir(parents=True, exist_ok=True)


def cell_val(cell):
    if cell is None:
        return None
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def sheet_to_rows(ws, max_rows=None, max_cols=None):
    rows = []
    rmax = ws.max_row if max_rows is None else min(ws.max_row, max_rows)
    cmax = ws.max_column if max_cols is None else min(ws.max_column, max_cols)
    for r in range(1, rmax + 1):
        row = []
        for c in range(1, cmax + 1):
            row.append(cell_val(ws.cell(row=r, column=c)))
        if any(v is not None for v in row):
            rows.append(row)
    return rows


def extract_workbook(path: Path, label: str):
    print(f"\n=== {label} ===")
    print(f"File: {path.name}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False, keep_vba=False)
    except Exception as e:
        print(f"  ERROR opening: {e}")
        return None

    out = {
        "file": path.name,
        "label": label,
        "sheets": {},
    }
    for sn in wb.sheetnames:
        try:
            ws = wb[sn]
        except Exception as e:
            print(f"  skip sheet {sn}: {e}")
            continue
        rows = sheet_to_rows(ws)
        out["sheets"][sn] = {
            "dim": [ws.max_row, ws.max_column],
            "rows": rows,
        }
        print(f"  [{sn}] {ws.max_row}r x {ws.max_column}c -> {len(rows)} non-empty")
    wb.close()
    return out


def main():
    targets = [
        ("pmp-latere",   "PMP LTE - Mantenimiento Programado 20240404.xlsm"),
        ("pmp-donchi",   "PMP DON CHICUETO - Mantenimiento Programado 20240404.xlsm"),
        ("plan-donchi-rev0", "PLAN DE MANTENIMIENTO DON CHICUETO REV00.xlsm"),
        ("plan-barcazas-2025", "1 Plan de mmto_Barcazas_Grupo de motores_2025.xlsx"),
        ("plan-bzas-prop", "PLAN DE MANTENIMIENTO BZAS_Rev 1_2025_propuesta.xlsx"),
        ("repuestos-criticos", "Lista de rptos criticos_motores de barcazas.xlsx"),
        ("analisis-repuestos", "Analisis de Respuestos.xlsx"),
        ("certificados", "1 Tablero de vencimientos documentos 01-01-2026.xlsx"),
        ("horas-motor", "Registro horas de motor de barcazas 2023.xlsx"),
        ("plan-barranqueras", "Plan de Mantenimiento MAQ. DICIEMBRE-2021 Barranqueras I.xlsx"),
    ]

    for key, fname in targets:
        path = ROOT / fname
        if not path.exists():
            print(f"MISSING: {fname}")
            continue
        data = extract_workbook(path, key)
        if data is None:
            continue
        out_path = OUT / f"{key}.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, default=str, indent=1)
        print(f"  -> {out_path.name}  ({out_path.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
