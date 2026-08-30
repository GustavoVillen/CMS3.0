# -*- coding: utf-8 -*-
"""Normaliza el PLAN DE MANT. LTE - JULIO.xlsx (PlanA) a JSON.

Detecta la fila de encabezado por el texto "Tarea a realizar" y de ahi saca los
indices de columna, porque cada hoja arranca en una columna distinta.
Arrastra el "Descripcion" (nombre del equipo) hacia abajo cuando viene vacio,
igual que la planilla en papel: el equipo se escribe una vez y las filas
siguientes son sus tareas.
"""
import json, re, sys, warnings, datetime
import openpyxl

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8")

SRC = "MisDocs/LTE/Mantenimiento/PLAN DE MANT. LTE - JULIO.xlsx"
SHEETS = [
    "MOTORES_PROPULSORES", "CAJAS_REDUCTORAS", "MOTORES_GENERADORES",
    "PLANTA_ELECTRICA", "CIRCUITO_DE_COMBUSTIBLE", "NAV-COM",
    "BOMBAS_ELECTRICAS", "COMPRESOR-A_A-TIFON", "SISTEMA_HIDRAULICO_PRINC_-AUX",
    "VENTILADORES,_EXTRACTORES-OTROS", "CABRESTANTE_DE_PROA_Y_POPA",
    "SEPARADOR_Y_PLANTA_P_T,E",
]
SKIP = re.compile(r"firma y aclaraci|^fecha:?$|^item$|^ítem$", re.I)


def norm(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip().replace("\n", " ")


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    out = []
    for sheet in SHEETS:
        ws = wb[sheet]
        rows = [[norm(c.value) for c in r] for r in ws.iter_rows(max_col=min(ws.max_column, 20))]
        # fila de encabezado
        hdr_i = next((i for i, r in enumerate(rows) if any("Tarea a realizar" in c for c in r)), None)
        if hdr_i is None:
            # NAV-COM no tiene fila de encabezado: A=nro, B=equipo, C=tarea,
            # D=frecuencia, E=ultima, F=proximo.
            hdr_i, c_task, c_desc, c_freq, c_last, c_next = 5, 2, 1, 3, 4, 5
        else:
            hdr = rows[hdr_i]
            c_task = next(i for i, c in enumerate(hdr) if "Tarea a realizar" in c)
            c_desc = c_task - 1
            c_freq = c_task + 1
            c_last = c_task + 2
            # "Hora de cambio Act." (motores) vs "Recorrido Actual" (resto)
            c_next = next((i for i, c in enumerate(hdr) if "recorrido" in c.lower() and "róximo" in c), c_task + 5)

        equipo = ""
        grupo = ""
        for i in range(hdr_i + 1, len(rows)):
            r = rows[i] + [""] * 20
            # cabecera repetida de pagina -> reset
            if any("Tarea a realizar" in c for c in r):
                continue
            task = r[c_task]
            # titulo de grupo: fila con texto en la 1a/2a columna y sin tarea
            first = next((c for c in r[:c_desc + 1] if c), "")
            if not task:
                if first and not SKIP.search(first):
                    grupo = first
                continue
            if SKIP.search(task):
                continue
            if r[c_desc]:
                equipo = r[c_desc]
            # grupo tipo "MOTOR PRINCIPAL N01 BR-KTA 50 CUMMINS" viene en col A
            if first and first != task and not r[c_desc] and len(first) > 8:
                grupo = first
            out.append({
                "sheet": sheet, "row": i + 1, "grupo": grupo, "equipo": equipo,
                "tarea": task, "frecuencia": r[c_freq],
                "ultima": r[c_last], "proximo": r[c_next],
            })
    json.dump(out, open("scripts/vessel-plan-tools/out/lte-plana.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print("filas:", len(out))
    for s in SHEETS:
        print(" ", s, sum(1 for o in out if o["sheet"] == s))


main()
