# -*- coding: utf-8 -*-
"""Escribe el Excel de la auditoria de planes de toda la flota.

Entrada: out/auditoria-flota.json (lo genera audit-fleet-plans.py)
Salida:  MisDocs/_Carga/Auditorias/<fecha>-Auditoria-Planes-Flota.xlsx

Cinco hojas: Guia, Resumen, Detalle, Diferencias y Sin mapear.

Uso: python scripts/vessel-plan-tools/report-fleet-audit-excel.py
"""
import datetime
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

ORIGEN = "scripts/vessel-plan-tools/out/auditoria-flota.json"
DESTINO_DIR = "MisDocs/_Carga/Auditorias"
HOY = datetime.date.today().isoformat()
DESTINO = os.path.join(DESTINO_DIR, "%s-Auditoria-Planes-Flota.xlsx" % HOY)

AZUL = "1F3B54"
GRIS = "F2F4F5"
COLORES = {                       # fondo de la celda de estado
    "EN AMBOS": "DCEAD9",
    "SOLO EXCEL": "F7D9D3",
    "SOLO CMS3": "FBEBCF",
    "RUTINA DE GUARDIA": "E3E7F2",
}
BORDE = Border(*[Side(style="thin", color="D6DCDF")] * 4)


def encabezar(ws, columnas, fila=1):
    for j, (titulo, ancho) in enumerate(columnas, start=1):
        c = ws.cell(row=fila, column=j, value=titulo)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = ancho
    ws.row_dimensions[fila].height = 30
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)
    ws.auto_filter.ref = "A%d:%s%d" % (fila, get_column_letter(len(columnas)), fila)


def escribir(ws, filas, campos, fila_inicial=2, color_estado=None):
    for i, f in enumerate(filas, start=fila_inicial):
        for j, campo in enumerate(campos, start=1):
            v = f.get(campo)
            if isinstance(v, bool):
                v = "SI" if v else ""
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=(j > 3))
            c.font = Font(size=10)
            c.border = BORDE
            if color_estado and campo == color_estado:
                col = COLORES.get(f.get(campo))
                if col:
                    c.fill = PatternFill("solid", fgColor=col)
                    c.font = Font(size=10, bold=True)


def hoja_guia(wb, datos):
    ws = wb.create_sheet("Guia")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 108
    ws["A1"] = "Auditoria de planes de mantenimiento — flota Mercurio"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells("A1:B1")

    lineas = [
        ("Que compara",
         "Las planillas Excel que manda el Owner, buque por buque, contra los planes de "
         "mantenimiento cargados en el CMS3."),
        ("Foto del sistema",
         "Dump de la base de produccion (VPS) del %s. No se modifico nada: es solo lectura."
         % datos["generado"][:16].replace("T", " ")),
        ("", ""),
        ("EN AMBOS",
         "La tarea de la planilla tiene su plan en el sistema. La columna 'Como se emparejo' "
         "dice por que se considera la misma tarea."),
        ("SOLO EXCEL",
         "Esta en la planilla del Owner y NO hay plan que la cubra. Es lo que falta cargar."),
        ("SOLO CMS3",
         "Hay un plan vivo en el sistema que la planilla no pide. Suele ser tarea de clase, "
         "del fabricante o heredada al clonar otro buque. No es necesariamente un error."),
        ("RUTINA DE GUARDIA",
         "Tarea diaria, semanal o quincenal de la ronda de la tripulacion. No genera plan "
         "propio: va en los planes consolidados de rutina. No cuenta como faltante."),
        ("", ""),
        ("Cobertura %",
         "Tareas EN AMBOS sobre el total de la planilla, sin contar las rutinas de guardia."),
        ("Motor (papel)",
         "Motor que declara la planilla de barcazas. 'Motor (sistema)' es la marca y modelo "
         "cargados en la ficha del activo. Si no coinciden, hay que revisarlo."),
        ("", ""),
        ("Limites de esta auditoria",
         "1) El emparejamiento es por activo y por texto de la tarea: dos tareas iguales "
         "escritas muy distinto pueden salir como SOLO EXCEL y SOLO CMS3 a la vez. "
         "2) Los equipos que la planilla nombra y no existen como activo en el sistema estan "
         "en la hoja 'Sin mapear'. "
         "3) Para las barcazas se uso el registro del SGS (REGI-MAN-02.2, rev. 30-abr-2026): "
         "conviene confirmar con Jorge que esa es la revision aprobada."),
    ]
    fila = 3
    for a, b in lineas:
        ws.cell(row=fila, column=1, value=a).font = Font(bold=True, size=10, color=AZUL)
        c = ws.cell(row=fila, column=2, value=b)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.font = Font(size=10)
        if a in COLORES:
            ws.cell(row=fila, column=1).fill = PatternFill("solid", fgColor=COLORES[a])
        ws.row_dimensions[fila].height = 15 if not b else max(15, 13 * (len(b) // 100 + 1))
        fila += 1


def hoja_resumen(wb, resumen):
    ws = wb.create_sheet("Resumen")
    cols = [("Buque", 10), ("Nombre", 18), ("Tipo", 22), ("Motor (papel)", 14),
            ("Tareas en la planilla", 12), ("Planes en el CMS3", 11),
            ("EN AMBOS", 10), ("SOLO EXCEL", 11), ("SOLO CMS3", 11),
            ("Rutina de guardia", 11), ("Cobertura %", 11), ("Con diferencias", 11)]
    encabezar(ws, cols)
    campos = ["buque", "buqueNombre", "tipo", "motorPapel", "papel", "sistema",
              "ambos", "soloExcel", "soloCms3", "rutina", "cobertura", "difieren"]
    escribir(ws, resumen, campos)

    # Fila de totales
    i = len(resumen) + 2
    ws.cell(row=i, column=1, value="TOTAL").font = Font(bold=True)
    for j, campo in enumerate(campos, start=1):
        if campo in ("papel", "sistema", "ambos", "soloExcel", "soloCms3", "rutina", "difieren"):
            c = ws.cell(row=i, column=j, value=sum(r[campo] for r in resumen))
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=GRIS)
    tot_papel = sum(r["papel"] for r in resumen)
    tot_ambos = sum(r["ambos"] for r in resumen)
    c = ws.cell(row=i, column=11, value=round(100.0 * tot_ambos / tot_papel, 1) if tot_papel else None)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor=GRIS)


def hoja_detalle(wb, filas):
    ws = wb.create_sheet("Detalle")
    cols = [("Estado", 18), ("Buque", 9), ("Nombre", 15), ("Activo", 18),
            ("Nombre del activo", 26), ("Codigo del plan", 20),
            ("Tarea segun la planilla", 52), ("Tarea segun el CMS3", 52),
            ("Frec. planilla", 12), ("Frec. CMS3", 12),
            ("Ult. ejec. planilla", 13), ("Ult. ejec. CMS3", 13),
            ("Vence planilla", 13), ("Vence CMS3", 13),
            ("Motor (papel)", 13), ("Motor (sistema)", 22),
            ("Equipo en la planilla", 26), ("Hoja de la planilla", 22), ("Fila", 7),
            ("OT", 6), ("Como se emparejo", 20), ("Diferencias", 46),
            ("Ejecutada despues del corte", 11)]
    encabezar(ws, cols)
    campos = ["estado", "buque", "buqueNombre", "activo", "activoNombre", "taskCode",
              "tituloPapel", "tituloSistema", "frecPapel", "frecSistema",
              "ultPapel", "ultSistema", "vencePapel", "venceSistema",
              "motorPapel", "motorSistema", "equipoPapel", "hoja", "fila",
              "ots", "via", "difiere", "adelantado"]
    orden = {"SOLO EXCEL": 0, "SOLO CMS3": 1, "EN AMBOS": 2, "RUTINA DE GUARDIA": 3}
    filas = sorted(filas, key=lambda f: (f["buque"], orden.get(f["estado"], 9),
                                         f["activo"] or "", f["tituloPapel"] or f["tituloSistema"] or ""))
    escribir(ws, filas, campos, color_estado="estado")


def hoja_diferencias(wb, filas):
    ws = wb.create_sheet("Diferencias")
    cols = [("Buque", 9), ("Activo", 18), ("Codigo del plan", 20),
            ("Tarea", 54), ("Que no coincide", 76),
            ("Ejecutada despues del corte", 12), ("OT", 6)]
    encabezar(ws, cols)
    campos = ["buque", "activo", "taskCode", "tituloPapel", "difiere", "adelantado", "ots"]
    d = [f for f in filas if f["estado"] == "EN AMBOS" and f.get("difiere")]
    d.sort(key=lambda f: (bool(f.get("adelantado")), f["buque"], f["activo"] or ""))
    escribir(ws, d, campos)
    return len(d)


def hoja_sinmapear(wb, sinmapear):
    ws = wb.create_sheet("Sin mapear")
    cols = [("Buque", 10), ("Equipo que nombra la planilla", 56), ("Mejor puntaje", 13),
            ("Que significa", 74)]
    encabezar(ws, cols)
    for f in sinmapear:
        f["nota"] = ("no se encontro un activo parecido en el sistema"
                     if not f["mejorPuntaje"] else
                     "hay un activo parecido pero no lo suficiente; revisar a mano")
    escribir(ws, sinmapear, ["buque", "equipo", "mejorPuntaje", "nota"])


def main():
    datos = json.load(open(ORIGEN, encoding="utf-8"))
    os.makedirs(DESTINO_DIR, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    hoja_guia(wb, datos)
    hoja_resumen(wb, datos["resumen"])
    hoja_detalle(wb, datos["filas"])
    n_dif = hoja_diferencias(wb, datos["filas"])
    hoja_sinmapear(wb, datos["sinMapear"])
    wb.save(DESTINO)

    print("Excel generado: %s" % DESTINO)
    print("   Resumen: %d buques" % len(datos["resumen"]))
    print("   Detalle: %d filas" % len(datos["filas"]))
    print("   Diferencias: %d" % n_dif)
    print("   Sin mapear: %d" % len(datos["sinMapear"]))


main()
